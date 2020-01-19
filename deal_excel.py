import openpyxl
# coding: UTF-8

def process_worksheet(sheet):
    avg_column = sheet.max_column + 1
    sum_column = sheet.max_column + 2
    for row in sheet.iter_rows(min_row = 2, min_col = 3):
        scores = [cell.value for cell in row]
        sum_score = sum(scores)
        avg_score = sum(scores)/len(scores)
        sheet.cell(row = row[0].row, column = avg_column).value = avg_score
        sheet.cell(row = row[0].row, column = sum_column).value = sum_score

        sheet.cell(row = 1, column = avg_column).value = "avg"
        sheet.cell(row = 1, column = sum_column).value = "sum"

def main():
    """docstring for main"""
    wb = openpyxl.load_workbook('example.xlsx')
    sheet = wb['student']
    process_worksheet(sheet)
    wb.save("example_copy.xlsx")

if __name__ == '__main__':
    main()

