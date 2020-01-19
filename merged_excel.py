import os
import openpyxl
import glob


def merge_xls_files(xlsx_files):
    """docstring for merge_xls_files"""
    wb = openpyxl.load_workbook(xlsx_files[0])
    ws = wb.active
    ws.title = 'merged result'
    for filename in xlsx_files[1:]:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            ws.append(value)
    return wb


def get_all_xlsx_files(path):
    """docstring for fname"""
    xlsx_files = glob.glob(os.path.join(path, '*.xlsx'))
    sorted(xlsx_files, key=str.lower)
    return xlsx_files


def main():
    """docstring for main"""
    xlsx_files = get_all_xlsx_files('XXX/XXX/xxx/')
    wb = merge_xlsx_files(xlsx_files)
    wb.save('merged_form.xlsx')


if __name__ == '__main__':
    main()
